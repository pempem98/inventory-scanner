import pandas as pd
import numpy as np
import os
import logging
import json
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import Dict, List, Optional, Tuple

# Cấu hình logging
logging.basicConfig(
    filename='runtime.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)

class ExcelSnapshotComparator:
    """So sánh hai file snapshot Excel và kiểm tra màu nền hợp lệ."""

    def __init__(
        self,
        file_predecessor: str,
        file_current: str,
        key_col: str,
        check_cols: List[str],
        allowed_key_pattern: str = r'^[A-Z0-9_.-]+$',
        invalid_colors: list[str] = ['#ff0000']
    ):
        """
        Khởi tạo với các file snapshot và cấu hình.

        Args:
            file_predecessor: Đường dẫn đến file predecessor Excel.
            file_current: Đường dẫn đến file current Excel.
            key_col: Tên cột khóa để so sánh (e.g., 'MÃ CĂN').
            check_cols: Danh sách các cột cần kiểm tra thay đổi.
            allowed_key_pattern: Regex pattern để kiểm tra giá trị khóa hợp lệ.
            invalid_colors: Danh sach các mã màu hex không hợp lệ (chuẩn hóa thành chữ thường).
        """
        self.file_predecessor = file_predecessor
        self.file_current = file_current
        self.key_col = key_col
        self.check_cols = check_cols
        self.allowed_key_pattern = allowed_key_pattern
        self.invalid_colors = invalid_colors

    def read_excel_data_and_colors(self, file_path: str) -> Tuple[pd.DataFrame, List[List[Optional[str]]]]:
        """Đọc dữ liệu và màu nền từ file Excel.

        Args:
            file_path: Đường dẫn đến file Excel.

        Returns:
            Tuple chứa DataFrame dữ liệu và danh sách màu nền (hex).

        Raises:
            Exception: Nếu lỗi khi đọc file.
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File {file_path} không tồn tại.")

            # Đọc dữ liệu
            df = pd.read_excel(file_path, header=None, engine='openpyxl')
            if df.empty:
                logging.warning(f"File {file_path} rỗng.")
                return df, []

            # Đọc màu nền
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            color_grid = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                row_colors = []
                for cell in row:
                    fill = cell.fill
                    color = None
                    if isinstance(fill, PatternFill) and fill.start_color and fill.start_color.rgb:
                        color = f"#{fill.start_color.rgb[2:].lower()}"  # Loại bỏ 'FF' prefix
                    row_colors.append(color)
                color_grid.append(row_colors)

            wb.close()
            return df, color_grid
        except Exception as e:
            logging.error(f"Lỗi khi đọc file Excel {file_path}: {e}")
            return pd.DataFrame(), []

    def normalize_key(self, key: str) -> str:
        """Chuẩn hóa giá trị khóa: giữ chữ cái tiếng Việt, số, dấu chấm, gạch dưới, gạch ngang; loại bỏ ký tự khác.

        Args:
            key: Giá trị khóa cần chuẩn hóa.

        Returns:
            Giá trị khóa đã chuẩn hóa.
        """
        if not isinstance(key, str):
            return ''
        # Regex giữ chữ cái tiếng Việt (bao gồm dấu), số, dấu chấm, gạch dưới, gạch ngang
        pattern = r'[A-Za-z0-9._-]+'
        normalized = ''.join(re.findall(pattern, key))
        # Loại bỏ khoảng trắng thừa
        normalized = re.sub(r'\s+', '', normalized)
        return normalized

    def compare(self) -> Dict[str, List]:
        """So sánh hai file snapshot và trả về các thay đổi, kiểm tra màu hợp lệ.

        Returns:
            Dictionary chứa danh sách các hàng thêm, xóa và thay đổi.
        """
        try:
            # Đọc file predecessor và current
            pred_df, pred_colors = self.read_excel_data_and_colors(self.file_predecessor)
            curr_df, curr_colors = self.read_excel_data_and_colors(self.file_current)

            if pred_df.empty or curr_df.empty:
                logging.warning(f"File rỗng: {self.file_predecessor} hoặc {self.file_current}")
                return {'added': [], 'removed': [], 'changed': [], 'remaining': []}

            # Tìm index của key_col và check_cols
            header = pred_df.iloc[0].tolist()
            if all(str(cell) == 'nan' for cell in header):
                header = [ chr(ord('A') + i) for i in range(len(header)) ]
            key_col_idx = header.index(self.key_col) if self.key_col in header else None
            if key_col_idx is None:
                logging.error(f"Không tìm thấy cột {self.key_col} trong {self.file_predecessor}")
                return {'added': [], 'removed': [], 'changed': [], 'remaining': []}

            check_col_indices = []
            for col in self.check_cols:
                if col in header:
                    check_col_indices.append(header.index(col))
                else:
                    logging.warning(f"Cột {col} không tồn tại trong {self.file_predecessor}")

            # Khởi tạo kết quả
            added = []
            removed = []
            changed = []
            remaining = []
            invalid_rows = []

            # Cột màu cho TÊN TÒA (cột 1 sau khi bỏ cột đầu tiên)
            color_col_idx = key_col_idx

            # Kiểm tra khóa hợp lệ
            key_pattern = re.compile(self.allowed_key_pattern)

            # Tìm các hàng bị xóa hoặc có màu thay đổi từ hợp lệ sang không hợp lệ
            for idx, row in pred_df.iterrows():
                key = self.normalize_key(row[key_col_idx])
                if (not key) or (len(key) >= 15) or (len(key) < 6) or not key_pattern.match(key):
                    logging.info(f"Bỏ qua hàng với khóa không hợp lệ: {key}")
                    continue

                curr_row_idx = curr_df[curr_df.iloc[:, key_col_idx].apply(self.normalize_key) == key].index
                if curr_row_idx.empty:
                    # Hàng không tồn tại trong file hiện tại
                    try:
                        cell_color = pred_colors[idx][color_col_idx]
                        if cell_color and cell_color.lower() in self.invalid_colors:
                            logging.info(f"Bỏ qua hàng với MÃ CĂN {key}: Màu {cell_color} không hợp lệ")
                            invalid_rows.append((key, cell_color))
                            continue
                        removed.append(key)
                    except IndexError:
                        logging.warning(f"Không tìm thấy màu cho hàng {idx+1}, cột {color_col_idx}")
                        continue
                else:
                    # Hàng tồn tại trong file hiện tại, kiểm tra màu
                    curr_row_idx = curr_row_idx[0]
                    try:
                        pred_cell_color = pred_colors[idx][color_col_idx]
                        curr_cell_color = curr_colors[curr_row_idx][color_col_idx]
                        if (pred_cell_color and pred_cell_color.lower() not in self.invalid_colors and
                            curr_cell_color and curr_cell_color.lower() in self.invalid_colors):
                            logging.info(f"Thêm vào removed: MÃ CĂN {key} vì màu thay đổi từ hợp lệ {pred_cell_color} sang không hợp lệ {curr_cell_color}")
                            removed.append(key)
                            invalid_rows.append((key, curr_cell_color))
                            continue
                    except IndexError:
                        logging.warning(f"Không tìm thấy màu cho hàng {idx+1} hoặc {curr_row_idx+1}, cột {color_col_idx}")
                        continue

            # Tìm các hàng được thêm hoặc còn lại
            for idx, row in curr_df.iterrows():
                key = self.normalize_key(row[key_col_idx])
                if (not key) or (len(key) >= 15) or (len(key) < 6) or not key_pattern.match(key):
                    logging.info(f"Bỏ qua hàng với khóa không hợp lệ: {key}")
                    continue

                try:
                    cell_color = curr_colors[idx][color_col_idx]
                    if cell_color and cell_color.lower() in self.invalid_colors:
                        logging.info(f"Bỏ qua hàng với MÃ CĂN {key}: Màu {cell_color} không hợp lệ")
                        invalid_rows.append((key, cell_color))
                        continue
                except IndexError:
                    logging.warning(f"Không tìm thấy màu cho hàng {idx+1}, cột {color_col_idx}")
                    continue

                if key not in pred_df.iloc[:, key_col_idx].apply(self.normalize_key).values:
                    # Hàng mới, thêm vào added
                    added.append(key)
                else:
                    # Hàng còn lại, thêm vào remaining nếu màu hợp lệ
                    remaining.append(key)

            # Tìm các hàng thay đổi
            for idx, pred_row in pred_df.iterrows():
                key = self.normalize_key(pred_row[key_col_idx])
                if not key or not key_pattern.match(key):
                    continue

                curr_row = curr_df[curr_df.iloc[:, key_col_idx].apply(self.normalize_key) == key]
                if not curr_row.empty:
                    curr_row = curr_row.iloc[0]
                    changes = {}
                    for col_idx in check_col_indices:
                        pred_val = pred_row[col_idx]
                        curr_val = curr_row[col_idx]
                        if pd.isna(pred_val) and pd.isna(curr_val):
                            continue
                        if pred_val != curr_val:
                            col_name = header[col_idx]
                            changes[col_name] = {'old': pred_val, 'new': curr_val}
                    if changes:
                        try:
                            cell_color = curr_colors[idx][color_col_idx]
                            if cell_color and cell_color.lower() in self.invalid_colors:
                                logging.info(f"Bỏ qua hàng với MÃ CĂN {key}: Màu {cell_color} không hợp lệ")
                                invalid_rows.append((key, cell_color))
                                continue
                        except IndexError:
                            logging.warning(f"Không tìm thấy màu cho hàng {idx+1}, cột {color_col_idx}")
                            continue
                        changed.append({'key': key, 'changes': changes})

            if invalid_rows:
                logging.info(f"Các hàng bị bỏ qua do màu không hợp lệ: {invalid_rows}")

            return {
                'added': added,
                'removed': removed,
                'changed': changed,
                'remaining': remaining
            }
        except Exception as e:
            logging.error(f"Lỗi khi so sánh {self.file_predecessor} và {self.file_current}: {e}")
            return {'added': [], 'removed': [], 'changed': [], 'remaining': []}